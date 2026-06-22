"""Выгрузка результатов опроса в Excel (.xlsx)."""

from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from bot.database import get_session
from bot.models import Direction, Participation


async def build_results_xlsx() -> BytesIO:
    """Один лист: по строке на каждое реальное (не тестовое) прохождение.

    Колонки: дата, имя, фамилия, username, Telegram ID, ведущее направление
    и по колонке с баллами на каждое направление.
    """
    async with get_session() as session:
        directions = list(
            await session.scalars(
                select(Direction).order_by(Direction.position, Direction.id)
            )
        )
        parts = list(
            await session.scalars(
                select(Participation)
                .where(Participation.is_test.is_(False))
                .options(
                    selectinload(Participation.respondent),
                    selectinload(Participation.top_direction),
                )
                .order_by(Participation.completed_at.desc())
            )
        )

    dir_names = [d.name for d in directions]
    headers = [
        "Дата",
        "Имя",
        "Фамилия",
        "Username",
        "Telegram ID",
        "Ведущее направление",
    ] + dir_names

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in parts:
        r = p.respondent
        scores = json.loads(p.scores_json or "{}")
        top = p.top_direction.name if p.top_direction else "—"
        row = [
            p.completed_at.strftime("%d.%m.%Y %H:%M") if p.completed_at else "",
            r.first_name or "",
            r.last_name or "",
            f"@{r.username}" if r.username else "",
            r.telegram_id,
            top,
        ] + [scores.get(name, 0) for name in dir_names]
        ws.append(row)

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
